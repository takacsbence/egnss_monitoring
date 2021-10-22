
import zipfile
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import math
from datetime import date
import os
import regex as re



#Ez az excel filebol beolvasott bazisallomas osztaly
class Base_station:
    def __init__(self, id, city, lat, long, elev):
        self.id = id
        self.city = city
        self.lat = lat
        self.long = long
        self.elev = elev

    def func(self):
        print("Ertekek" + self.id + self.city)

#Ez a pos filebol olvasott adatok osztalya
class Gps_data:
    def __init__(self, date, time, lat, lon, ele, mode, nsat, stdn, stde, stdu, stdne, stdeu, stdun, age, ratio):
        self.date = date
        self.time = time
        self.lat = lat
        self.lon = lon
        self.ele = ele
        self.mode = mode
        self.nsat = nsat
        self.stdn = stdn
        self.stde = stde
        self.stdu = stdu
        self.stdne = stdne
        self.stdeu = stdeu
        self.sttun = stdun
        self.age = age
        self.ratio = ratio

#FTP request majd ide


#A majd ftps szerverrol letoltott file kicsomagolasa
zip_path = "D:\Rinex_datas\HC-Nyiregyhaza_17-10-2021_17-10-2021.zip"
zip_extract_path = "D:\Rinex_datas"

with zipfile.ZipFile(zip_path, 'r') as zip_ref:
    zip_ref.extractall(zip_extract_path)

#Sikerult mukodesre birni az rnx2rtp-t is ebben a formaban
fp="D:\Rinex_datas\\rnx2rtkp -p 0 -f 1 -f 2 -f 5 -t -e D:\Rinex_datas\HC-Budapest_15-10-2021_15-10-2021.o D:\Rinex_datas\HC-Budapest_15-10-2021_15-10-2021.n -o D:\Rinex_datas\HC-Budapest_15-10-2021_15-10-2021.pos"
os.system(fp)

#a bazisallomasok tarolasa
listofbst = []
#a pos filebol beolvasott adatok tarolasa
listofpositions = []

#a felhasznalo megadja azt hogy melyik idotartamra szeretne kiertekelni a pos filet
u_input=input("Melyik órára szeretnéd megvizsgálni?")

#ket valtozo ezek lesznek beallitva a megadott idotartam szerint a megfelelo ertekre, hogy a pos file soraibol a jo adatot olvassa az idonek megfeleloen
start=0
stop=0

if(u_input=='0'):
    start = 16
    stop = 3616
elif(u_input=='1'):
    start = 3616
    stop = 7216
elif(u_input == '2'):
    start = 7216
    stop = 10816
elif (u_input == '3'):
    start = 10816
    stop = 14416
elif (u_input == '4'):
    start = 14416
    stop = 18016
elif (u_input == '5'):
    start = 18016
    stop = 21616
elif (u_input == '6'):
    start = 21616
    stop = 25216
elif (u_input == '7'):
    start = 25216
    stop =28816
elif (u_input == '8'):
    start=28816
    stop=32416
elif (u_input == '9'):
    start=32416
    stop=36016
elif (u_input == '10'):
    start=36016
    stop=39616
elif (u_input == '11'):
    start=39616
    stop=43216
elif (u_input == '12'):
    start=43216
    stop=46816
elif (u_input == '13'):
    start=46816
    stop=50416
elif (u_input == '14'):
    start=50416
    stop=54016
elif (u_input == '15'):
    start=54016
    stop=57616
elif (u_input == '16'):
    start=57616
    stop=61216
elif (u_input == '17'):
    start=61216
    stop=64816
elif (u_input == '18'):
    start=64816
    stop=68416
elif (u_input == '19'):
    start=68416
    stop=72016
elif (u_input == 19):
    time_stamps = 19
elif (u_input == 20):
    time_stamps = 20
elif (u_input == 21):
    time_stamps = 21
elif (u_input == 22):
    time_stamps = 22


path_to_pos = "D:\Rinex_datas\HC-Nyiregyhaza_17-10-2021_17-10-2021.pos"


#Ez a valtozo tarolja, hogy melyik bazisallomashoz hasonlitsuk a pos filet, ugy hogy ellenorzi a pos file nevet
base_station=0
station_name=""

if(path_to_pos.__contains__("Budapest")):
    station_name="Budapest"
    base_station=0
elif(path_to_pos.__contains__("Nyiregyhaza")):
    station_name="Nyiregyhaza"
    base_station=1
elif(path_to_pos.__contains__("Sarmellek")):
    station_name = "Sarmellek"
    base_station=2
elif(path_to_pos.__contains__("Szeged")):
    station_name = "Szeged"
    base_station=3
elif(path_to_pos.__contains__("Gyor-Per")):
    station_name = "Gyor-Per"
    base_station=4
elif(path_to_pos.__contains__("Bekescsaba")):
    station_name = "Bekescsaba"
    base_station=5
elif(path_to_pos.__contains__("Debrecen")):
    station_name = "Debrecen"
    base_station=6
elif(path_to_pos.__contains__("Pecs-Pogany")):
    station_name = "Pecs-Pogany"
    base_station=7
elif(path_to_pos.__contains__("Bugac")):
    station_name = "Bugac"
    base_station=8
elif(path_to_pos.__contains__("Sajohidveg")):
    station_name = "Sajohidveg"
    base_station=9
elif(path_to_pos.__contains__("Sagvar")):
    station_name = "Sagvar"
    base_station=10


#Itt olvassa be a pos filet jelenleg a felhasznalo altal kivalasztott idokereten belul
f = open("D:\Rinex_datas\HC-Nyiregyhaza_17-10-2021_17-10-2021.pos",'r')
for x, line in enumerate(f):
    if x >= start and x <= stop:
        converted = str(line)
        splitted = converted.split()
        element=Gps_data(splitted[0],splitted[1],float(splitted[2]),float(splitted[3]),float(splitted[4]),splitted[5],splitted[6],splitted[7],splitted[8],splitted[9],splitted[10],splitted[11],splitted[12],splitted[13],splitted[14])
        listofpositions.append(element)

difference_list=[]

path = "D:\Rinex_datas\stations4.xlsx"
wb = load_workbook(path)
ws = wb.active
x = 2

#Itt olvassa be a bazisallomasokat az excel filebol
while x != 12:
    converter = str(x)
    ID = ws['A' + converter].value
    city = ws['B' + converter].value
    latitude = ws['C' + converter].value
    longitude = ws['D' + converter].value
    elevation = ws['E' + converter].value
    basestc = Base_station(ID, city, latitude, longitude, elevation)
    listofbst.append(basestc)
    x = x + 1

y=0
x=0
Rfold=6378.137
rad=0.000008998719243599958
differences_by_city=[]

#Itt tortenik az elteres szamolasa, a fok kulonbseg jonak tunik de a meterre valo atvaltas rossz
while y != len(listofpositions):
    diff_in_degree=listofbst[base_station].long - listofpositions[y].lon
    dlat=(listofbst[base_station].lat*math.pi/180) - (listofpositions[y].lat*math.pi/180)
    print("dlat:",str(dlat))
    dlong=(listofbst[base_station].long*math.pi/180) - (listofpositions[y].lon*math.pi/180)
    print("dlon", str(dlong))
    a_val=math.sin(dlat/2)*math.sin(dlat/2)+ math.cos((listofbst[base_station].lat)*math.pi/180)*math.cos((listofpositions[y].lat)*math.pi/180)*math.sin(dlong/2)*math.sin(dlong/2)
    c=2*math.atan2(math.sqrt(a_val),math.sqrt(1-a_val))
    diff_in_kms=Rfold*c
    diff_in_meters=diff_in_kms*1000
    val=(math.sqrt(math.pow(listofbst[base_station].lat-listofpositions[y].lat,2))+math.pow(listofbst[base_station].long-listofpositions[y].lon,2))/rad
    print("d in meters:",str(diff_in_meters))
    print("val:",str(val))
    conv_to_meter=diff_in_degree*31
    differences_by_city.append(diff_in_meters)
    y=y+1

x_axis=[]
y_axis=[]
i=0
#A grafikon x tengelyet feltolti 3600s-al
while i != 3600:
    x_axis.append(float(i))
    i=i+1

#Az y tengely pedig az eltereseket abrazolja
y_axis=differences_by_city

today = date.today()

#Itt tortenik a grafikon kirajzolasa
it=0
while it!=3600:
    plt.scatter(x_axis[it], y_axis[it], color="red",marker="*", s=30)
    it=it+1

plt.xlabel('x - axis in seconds')
plt.ylabel('y - axis in meters')
plt.title('E-W Pos differences')
plt.legend()

#Itt menti el egy mappaba a grafikonokat kepkent az adott bazisallomas nevevel es az adott nap datumaval
plt.savefig("D:\Rinex_datas\E-W_graphs\\"+station_name+str(u_input)+"ora"+".png")
plt.show()


