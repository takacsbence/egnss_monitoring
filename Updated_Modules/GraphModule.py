import os.path
import sys

from matplotlib.dates import DateFormatter
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd

#Ez az excel filebol beolvasott bazisallomas osztaly
class Base_station:
    def __init__(self, id, city, lat, long, elev):
        self.id = id
        self.city = city
        self.lat = lat
        self.long = long
        self.elev = elev

    def func(self):
        print("Ertekek" + self.id + self.city)

#Ez a pos filebol olvasott adatok osztalya
class Gps_data:
    def __init__(self, date, time, lat, lon, ele, mode, nsat, stdn, stde, stdu, stdne, stdeu, stdun, age, ratio):
        self.date = date
        self.time = time
        self.lat = lat
        self.lon = lon
        self.ele = ele
        self.mode = mode
        self.nsat = nsat
        self.stdn = stdn
        self.stde = stde
        self.stdu = stdu
        self.stdne = stdne
        self.stdeu = stdeu
        self.sttun = stdun
        self.age = age
        self.ratio = ratio

#check number of args
if len(sys.argv) != 6:
    print("arguments: yyy ddd sss h")
    print("where: yyy  y=year, ddd=day of year, sss=station id, h=hourly session with abc, uid=user id")
    print("an example 2021 10 205 a John")
    exit()

#Recommended user inputs
year=str(sys.argv[1])

doy=str(sys.argv[2])

station=str(sys.argv[3])

time=str(sys.argv[4])

uid=str(sys.argv[5])




#Use of config file to determine save location
temp=""
pic_save=""
file = open('config.txt', 'r')
for line in file:
        if(line.split(' ')[0]==sys.argv[5]):
            temp=line.split(' ')[1]
            pic_save=line.split(' ')[2]

file.close()

save_location=temp[6:]
pic_save=pic_save[10:-1]

#Compare station id to station name
base_station=0
station_name=""

if(station=='205'):
    station_name="Budapest"
    base_station=0
elif(station=='206'):
    station_name="Nyiregyhaza"
    base_station=1
elif(station=='207'):
    station_name = "Sarmellek"
    base_station=2
elif(station=='208'):
    station_name = "Szeged"
    base_station=3
elif(station=='209'):
    station_name = "Gyor-Per"
    base_station=4
elif(station=='211'):
    station_name = "Bekescsaba"
    base_station=5
elif(station=='212'):
    station_name = "Debrecen"
    base_station=6
elif(station=='213'):
    station_name = "Pecs-Pogany"
    base_station=7
elif(station=='214'):
    station_name = "Bugac"
    base_station=8
elif(station=='215'):
    station_name = "Sajohidveg"
    base_station=9
elif(station=='210'):
    station_name = "Sagvar"
    base_station=10

year_for_filename= year[2:]
station='PildoBox'+station
file_name=save_location +'//' + station + year_for_filename + doy + time + '.pos'

if not os.path.exists(file_name):
    print('Position file does not exists!')
    exit()

dupe_check= pic_save +'//' + station +"_"+ year+"_" + doy+"_" + time + '.png'


path = "stations4.xlsx"
wb = load_workbook(path)
ws = wb.active
x = 2

#Store
listofbst = []


#Read station data from excel file
while x != 12:
    converter = str(x)
    ID = ws['A' + converter].value
    city = ws['B' + converter].value
    latitude = ws['C' + converter].value
    longitude = ws['D' + converter].value
    elevation = ws['E' + converter].value
    basestc = Base_station(ID, city, latitude, longitude, elevation)
    listofbst.append(basestc)
    x = x + 1

##Todo Dupe_Check
if os.path.exists(dupe_check):
    print("Duplicate Detected!")
    exit()

#Read position file
data_gps=pd.read_csv(file_name,header=None,delim_whitespace=True,skiprows=15)
data_gps.columns=["date", "time", "lat", "lon", "ele", "mode", "nsat", "stdn", "stde", "stdu", "stdne", "stdeu", "stdun", "age", "ratio",]

#Calculate position error in meters
EW_meters=(data_gps['lon']-listofbst[base_station].long)*21*3600
NS_meters=(data_gps['lat']-listofbst[base_station].lat)*31*3600
Elevation=data_gps['ele']-listofbst[base_station].elev

if not os.path.exists(pic_save):
    os.mkdir(pic_save)

data_gps['datetime'] = pd.to_datetime(data_gps['date'] + ' ' + data_gps['time'], format='%Y/%m/%d %H:%M:%S.%f')
fig, ax = plt.subplots(2)
fig.tight_layout()
ax[0].plot(data_gps['datetime'], EW_meters)

#EW-graph plot
ax[0].set_ylim([-1.5, 1.5])
ax[0].set_xlim([min(data_gps['datetime']).round('60min').to_pydatetime(), max(data_gps['datetime']).round('60min').to_pydatetime()]) #show exactly one hour session
ax[0].set_xlabel('time (hh:mm)')
ax[0].set_ylabel('Difference in meters')
ax[0].xaxis.set_major_formatter(DateFormatter("%H:%M"))
ax[0].grid()
ax[0].set_title('EW-Position')
#plt.savefig(pic_save+'//' + station_name+"_" +year+"_"+doy+"_"+ time + ".png")

#NS-graph plot
ax[1].plot(data_gps['datetime'], NS_meters)
ax[1].set_ylim([-4,4])
ax[1].set_xlim([min(data_gps['datetime']).round('60min').to_pydatetime(), max(data_gps['datetime']).round('60min').to_pydatetime()]) #show exactly one hour session
ax[1].set_xlabel('time (hh:mm)')
ax[1].set_ylabel('Difference in meters')
ax[1].xaxis.set_major_formatter(DateFormatter("%H:%M"))
ax[1].grid()
ax[1].set_title('NS-Position')
fig.set_size_inches(10, 10)
plt.savefig(pic_save+'//EW_NS_pos_' + station_name+"_" +year+"_"+doy+"_"+ time + ".png",dpi=100)

#Elevation-graph plot
fig2, ax2 = plt.subplots(2)
ax2[0].plot(data_gps['datetime'],Elevation )
ax2[0].set_ylim([-10, 10])
ax2[0].set_xlim([min(data_gps['datetime']).round('60min').to_pydatetime(), max(data_gps['datetime']).round('60min').to_pydatetime()]) #show exactly one hour session
ax2[0].set_xlabel('time (hh:mm)')
ax2[0].set_ylabel('Difference in meters')
ax2[0].xaxis.set_major_formatter(DateFormatter("%H:%M"))
ax2[0].grid()
ax2[0].set_title('Elevation Difference')
#plt.savefig(pic_save+'//' + station_name+"_" +year+"_"+doy+"_"+ time + ".png",dpi=100)

#Number of satellites
ax2[1].plot(data_gps['datetime'],data_gps['nsat'])
ax2[1].set_ylim([0, 15])
ax2[1].set_xlim([min(data_gps['datetime']).round('60min').to_pydatetime(), max(data_gps['datetime']).round('60min').to_pydatetime()]) #show exactly one hour session
ax2[1].set_xlabel('time (hh:mm)')
ax2[1].set_ylabel('Number of available satellites')
ax2[1].xaxis.set_major_formatter(DateFormatter("%H:%M"))
ax2[1].grid()
ax2[1].set_title('Number of Satellites')
fig2.set_size_inches(10, 10)
plt.savefig(pic_save+'//Elev_Nsat_' + station_name+"_" +year+"_"+doy+"_"+ time + ".png",dpi=100)

print('Graphs created and saved!')

#plt.show()

zip_path= save_location +'\\PildoBox' + station + year_for_filename + doy + time + ".raw.zip"

for f in os.listdir(save_location):
    if f.endswith('.nav') or f.endswith('.lnav') or f.endswith('.hnav') or f.endswith('.obs') or f.endswith('.pos') or f.endswith('.raw') or f.endswith('.stat') or f.endswith('.sbs'):
        #print("torolt file:\n",str(f))
        os.remove(save_location+"//"+f)



